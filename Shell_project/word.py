import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_workbook
from openpyxl.formula.translate import Translator

# Create a new workbook and add a sheet
wb = Workbook()
ws = wb.active
ws.title = "Investment Tracker"

# Headers for the tracker
headers = ["Month", "GoldBees Investment", "GoldBees Value",
           "PPF Investment", "PPF Value",
           "Nifty 50 Investment", "Nifty 50 Value"]

# Add headers to sheet
ws.append(headers)

# Setup for 15 years = 180 months
num_months = 180

# Populate the first column with months and leave investments blank (user input)
for month in range(1, num_months + 1):
    ws.cell(row=month + 1, column=1).value = f"Month {month}"
    # GoldBees, PPF, Nifty investment columns (initially blank for user to fill)
    for col in [2, 4, 6]:
        ws.cell(row=month + 1, column=col).value = None

# Define interest rates per month (for automatic compounding):
# GoldBees assumed annual rate: 10% => monthly: 10/12
# PPF assumed annual rate: 7.1% => monthly: 7.1/12
# Nifty 50 Index assumed annual rate: 12% => monthly: 12/12

monthly_rates = {
    "GoldBees": 10 / 12 / 100,
    "PPF": 7.1 / 12 / 100,
    "Nifty": 12 / 12 / 100
}

# Add formulas for calculating value with compounding
for row in range(2, num_months + 2):
    # For GoldBees (columns C: value = prev_value * (1+rate) + investment)
    if row == 2:
        ws.cell(row=row, column=3).value = f"=B{row}*(1+{monthly_rates['GoldBees']})"
    else:
        ws.cell(row=row, column=3).value = f"=C{row-1}*(1+{monthly_rates['GoldBees']})+B{row}"

    # For PPF (columns E)
    if row == 2:
        ws.cell(row=row, column=5).value = f"=D{row}*(1+{monthly_rates['PPF']})"
    else:
        ws.cell(row=row, column=5).value = f"=E{row-1}*(1+{monthly_rates['PPF']})+D{row}"

    # For Nifty (columns G)
    if row == 2:
        ws.cell(row=row, column=7).value = f"=F{row}*(1+{monthly_rates['Nifty']})"
    else:
        ws.cell(row=row, column=7).value = f"=G{row-1}*(1+{monthly_rates['Nifty']})+F{row}"

# Save the workbook
file_path = "/workspaces/Devops/Blank_Monthly_Investment_AutoCalc.xlsx"
wb.save(file_path)

file_path
